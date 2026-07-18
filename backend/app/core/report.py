"""Report mensili Excel (Fase 2, M11): proiezione deterministica delle viste.

"Il cliente continua a vivere in Excel senza doverlo abbandonare" (§3.8): i
report standard si generano come ``.xlsx`` dalle viste DuckDB, senza LLM. Un
foglio per entità (fatture, DDT, ore, SAL, scostamento) più un riepilogo per
cantiere. Filtrabile per cantiere e per mese.
"""

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.core.views import query

EURO = '#,##0.00 "€"'
PERC = "0.0%"
FONT_INTEST = Font(bold=True, color="FFFFFF")
FILL_INTEST = PatternFill("solid", fgColor="334155")
FONT_TOTALE = Font(bold=True)

MESI = [
    "", "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
    "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre",
]


def _filtro(
    cantiere_id: str | None,
    anno: int | None,
    mese: int | None,
    col_data: str = "data",
    col_cant: str = "cantiere_id",
) -> tuple[str, list[Any]]:
    """Clausola WHERE e parametri per cantiere e/o mese (col_data=None: no data)."""
    clausole: list[str] = []
    parametri: list[Any] = []
    if cantiere_id:
        clausole.append(f"{col_cant} = ?")
        parametri.append(cantiere_id)
    if col_data and anno:
        clausole.append(f"year({col_data}) = ?")
        parametri.append(anno)
    if col_data and mese:
        clausole.append(f"month({col_data}) = ?")
        parametri.append(mese)
    where = f" WHERE {' AND '.join(clausole)}" if clausole else ""
    return where, parametri


# ---- colonne dei fogli: (intestazione, chiave, formato) -------------------

Colonna = tuple[str, str, str | None]


def _foglio(
    wb: Workbook,
    titolo: str,
    colonne: list[Colonna],
    righe: list[dict[str, Any]],
    totali: set[str] = frozenset(),
) -> Worksheet:
    ws = wb.create_sheet(titolo)
    ws.append([c[0] for c in colonne])
    for cella in ws[1]:
        cella.font = FONT_INTEST
        cella.fill = FILL_INTEST
    for riga in righe:
        ws.append([riga.get(c[1]) for c in colonne])
    for indice, (_, _chiave, formato) in enumerate(colonne, start=1):
        if formato:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=indice).number_format = formato
    if totali and righe:
        _riga_totale(ws, colonne, righe, totali)
    _larghezze(ws, colonne)
    return ws


def _riga_totale(
    ws: Worksheet, colonne: list[Colonna], righe: list[dict[str, Any]], totali: set[str]
) -> None:
    valori: list[Any] = []
    for _, chiave, _fmt in colonne:
        valori.append(sum(r.get(chiave) or 0 for r in righe) if chiave in totali else None)
    if valori:
        valori[0] = "TOTALE"
    ws.append(valori)
    for indice, (_, _chiave, formato) in enumerate(colonne, start=1):
        cella = ws.cell(row=ws.max_row, column=indice)
        cella.font = FONT_TOTALE
        if formato:
            cella.number_format = formato


def _larghezze(ws: Worksheet, colonne: list[Colonna]) -> None:
    for indice, (intestazione, _chiave, _fmt) in enumerate(colonne, start=1):
        massimo = len(str(intestazione))
        for r in range(2, ws.max_row + 1):
            valore = ws.cell(row=r, column=indice).value
            massimo = max(massimo, len(str(valore)) if valore is not None else 0)
        lettera = ws.cell(row=1, column=indice).column_letter
        ws.column_dimensions[lettera].width = min(48, massimo + 3)


# ---- costruzione del workbook ---------------------------------------------


def _riepilogo(data_dir: Path, cantiere_id: str | None, anno: int | None, mese: int | None):
    where_f, pf = _filtro(cantiere_id, anno, mese, col_cant="cantiere_id")
    spese = {
        r["cantiere_id"]: r
        for r in query(
            data_dir,
            f"SELECT cantiere_id, COUNT(*) AS n, COALESCE(SUM(totale),0) AS speso "
            f"FROM v_fatture{where_f} GROUP BY cantiere_id",
            pf,
        )
    }
    ore = {
        r["cantiere_id"]: r
        for r in query(
            data_dir,
            f"SELECT cantiere_id, COALESCE(SUM(ore),0) AS ore, COALESCE(SUM(costo),0) AS costo "
            f"FROM v_rapportini_righe{where_f} GROUP BY cantiere_id",
            list(pf),
        )
    }
    cant_where = " WHERE id = ?" if cantiere_id else ""
    cantieri = query(
        data_dir,
        f"SELECT id, nome, budget FROM v_cantieri{cant_where} ORDER BY nome",
        [cantiere_id] if cantiere_id else [],
    )
    return [
        {
            "cantiere": c["nome"],
            "budget": c["budget"],
            "speso": spese.get(c["id"], {}).get("speso", 0),
            "n_fatture": spese.get(c["id"], {}).get("n", 0),
            "ore": ore.get(c["id"], {}).get("ore", 0),
            "costo_manodopera": ore.get(c["id"], {}).get("costo", 0),
        }
        for c in cantieri
    ]


def _righe(data_dir: Path, sql: str, cantiere_id, anno, mese, **kw) -> list[dict[str, Any]]:
    where, params = _filtro(cantiere_id, anno, mese, **kw)
    return query(data_dir, sql.format(where=where), params)


def titolo_periodo(anno: int | None, mese: int | None) -> str:
    if anno and mese:
        return f"{MESI[mese]} {anno}"
    if anno:
        return str(anno)
    return "tutto il periodo"


def genera_report(
    data_dir: Path | str,
    cantiere_id: str | None = None,
    anno: int | None = None,
    mese: int | None = None,
) -> bytes:
    """Costruisce il report .xlsx e ne ritorna i byte."""
    data_dir = Path(data_dir)
    wb = Workbook()
    wb.remove(wb.active)

    copertina = wb.create_sheet("Riepilogo")
    copertina.append(["Report Workflower", titolo_periodo(anno, mese)])
    copertina["A1"].font = Font(bold=True, size=14)
    copertina.append([])
    _blocco_riepilogo(wb, _riepilogo(data_dir, cantiere_id, anno, mese))

    _foglio(wb, "Fatture", _COL_FATTURE, _righe(
        data_dir,
        "SELECT f.id, f.numero, f.data, fo.ragione_sociale AS fornitore, f.cantiere_id, "
        "f.imponibile, f.iva, f.totale, f.ritenuta_acconto, f.stato "
        "FROM v_fatture f LEFT JOIN v_fornitori fo ON fo.id = f.fornitore_id{where} "
        "ORDER BY f.data DESC",
        cantiere_id, anno, mese, col_data="f.data", col_cant="f.cantiere_id",
    ), totali={"imponibile", "iva", "totale"})

    _foglio(wb, "DDT", _COL_DDT, _righe(
        data_dir,
        "SELECT d.id, d.numero, d.data, fo.ragione_sociale AS fornitore, d.cantiere_id, "
        "d.causale, d.n_righe, d.stato "
        "FROM v_ddt d LEFT JOIN v_fornitori fo ON fo.id = d.fornitore_id{where} "
        "ORDER BY d.data DESC",
        cantiere_id, anno, mese, col_data="d.data", col_cant="d.cantiere_id",
    ))

    _foglio(wb, "Ore", _COL_ORE, _righe(
        data_dir,
        "SELECT cantiere_id, data, nominativo, mansione, ore, costo_orario, costo "
        "FROM v_rapportini_righe{where} ORDER BY data DESC, nominativo",
        cantiere_id, anno, mese,
    ), totali={"ore", "costo"})

    _foglio(wb, "SAL", _COL_SAL, _righe(
        data_dir,
        "SELECT cantiere_id, numero, data, importo_lavori, importo_progressivo, "
        "percentuale_avanzamento, stato FROM v_sal{where} ORDER BY data DESC",
        cantiere_id, anno, mese,
    ))

    _foglio(wb, "Scostamento computo", _COL_SCOST, _righe(
        data_dir,
        "SELECT cantiere_id, codice, descrizione, categoria, previsto, consuntivo, delta, quota "
        "FROM v_scostamento_voci{where} ORDER BY cantiere_id, consuntivo DESC",
        cantiere_id, None, None, col_data=None,
    ), totali={"previsto", "consuntivo"})

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _blocco_riepilogo(wb: Workbook, righe: list[dict[str, Any]]) -> None:
    """Scrive la tabella di riepilogo nel foglio già creato (dopo la copertina)."""
    ws = wb["Riepilogo"]
    intestazioni = ["Cantiere", "Budget", "Speso fatture", "N. fatture", "Ore", "Costo manodopera"]
    riga_intest = ws.max_row + 1
    ws.append(intestazioni)
    for cella in ws[riga_intest]:
        cella.font = FONT_INTEST
        cella.fill = FILL_INTEST
    chiavi = ["cantiere", "budget", "speso", "n_fatture", "ore", "costo_manodopera"]
    for r in righe:
        ws.append([r[k] for k in chiavi])
    for r in range(riga_intest + 1, ws.max_row + 1):
        for col in (2, 3, 6):
            ws.cell(row=r, column=col).number_format = EURO
    ws["A1"].alignment = Alignment(horizontal="left")
    for col, larghezza in (("A", 34), ("B", 16), ("C", 16), ("D", 12), ("E", 10), ("F", 18)):
        ws.column_dimensions[col].width = larghezza


_COL_FATTURE: list[Colonna] = [
    ("ID", "id", None), ("Numero", "numero", None), ("Data", "data", None),
    ("Fornitore", "fornitore", None), ("Cantiere", "cantiere_id", None),
    ("Imponibile", "imponibile", EURO), ("IVA", "iva", EURO), ("Totale", "totale", EURO),
    ("Ritenuta", "ritenuta_acconto", EURO), ("Stato", "stato", None),
]
_COL_DDT: list[Colonna] = [
    ("ID", "id", None), ("Numero", "numero", None), ("Data", "data", None),
    ("Fornitore", "fornitore", None), ("Cantiere", "cantiere_id", None),
    ("Causale", "causale", None), ("Righe", "n_righe", None), ("Stato", "stato", None),
]
_COL_ORE: list[Colonna] = [
    ("Cantiere", "cantiere_id", None), ("Data", "data", None), ("Nominativo", "nominativo", None),
    ("Mansione", "mansione", None), ("Ore", "ore", None), ("Costo orario", "costo_orario", EURO),
    ("Costo", "costo", EURO),
]
_COL_SAL: list[Colonna] = [
    ("Cantiere", "cantiere_id", None), ("Numero", "numero", None), ("Data", "data", None),
    ("Importo lavori", "importo_lavori", EURO), ("Progressivo", "importo_progressivo", EURO),
    ("Avanzamento %", "percentuale_avanzamento", None), ("Stato", "stato", None),
]
_COL_SCOST: list[Colonna] = [
    ("Cantiere", "cantiere_id", None), ("Codice", "codice", None),
    ("Descrizione", "descrizione", None), ("Categoria", "categoria", None),
    ("Previsto", "previsto", EURO), ("Consuntivo", "consuntivo", EURO),
    ("Delta", "delta", EURO), ("Consumo", "quota", PERC),
]
