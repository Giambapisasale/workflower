"""Simulazione end-to-end: un mese di attività su 10 cantieri, 100 dipendenti.

Popola un repo dati con un mese realistico (88 operai di varie mansioni + 12
impiegati d'ufficio, rapportini giornalieri, DDT, fatture, SAL, pozzetti,
cronoprogrammi, scadenze) e verifica che tutte le superfici di lettura della
piattaforma — viste, cruscotto, registro di cantiere, scostamenti, report Excel,
RBAC — siano coerenti con gli aggregati calcolati in modo indipendente.

Lo scenario è costruito una volta sola (fixture di sessione) e interrogato in
sola lettura da molti test.
"""

import datetime as dt
import io

import pytest
from fake_llm import FakeCompleter
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from simulazione import costruisci_mese

from app.core.gateway import Gateway
from app.core.views import query
from app.main import create_app

TOLL = 0.01


@pytest.fixture(scope="session")
def mese(tmp_path_factory: pytest.TempPathFactory):
    data_dir = tmp_path_factory.mktemp("simulazione") / "data"
    scenario = costruisci_mese(data_dir)
    return data_dir, scenario


@pytest.fixture
def client_mese(mese) -> TestClient:
    data_dir, _ = mese
    gateway = Gateway(completer=FakeCompleter(data_dir), attesa_retry=0)
    return TestClient(create_app(data_dir=data_dir, gateway=gateway))


def _accedi(client: TestClient, username: str, pin: str) -> dict[str, str]:
    r = client.post("/api/auth/login", json={"username": username, "pin": pin})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['token']}"}


# ---------------------------------------------------------------- scala


def test_scala_dello_scenario(mese) -> None:
    _data_dir, sc = mese
    assert len(sc.cantieri) == 10
    assert sc.n_rapportini == 220          # 10 cantieri × 22 giorni lavorativi
    assert sc.n_fatture == 50 and sc.n_ddt == 80 and sc.n_sal == 20
    assert len(sc.operai) == 88 and len(sc.impiegati) == 12
    assert len(sc.operai) + len(sc.impiegati) == 100   # organico totale
    assert len(sc.mansioni) >= 8                        # operai di vario tipo


def test_organico_nelle_viste(mese) -> None:
    data_dir, sc = mese
    distinti = query(data_dir, "SELECT COUNT(DISTINCT nominativo) AS n FROM v_rapportini_righe")
    assert distinti[0]["n"] == 88   # ogni operaio compare almeno una volta nel mese
    mansioni = query(data_dir, "SELECT COUNT(DISTINCT mansione) AS n FROM v_rapportini_righe")
    assert mansioni[0]["n"] == len(sc.mansioni)


# ---------------------------------------------------------------- integrità


def test_ogni_fattura_ha_totale_coerente(mese) -> None:
    data_dir, _ = mese
    incoerenti = query(
        data_dir,
        "SELECT id FROM v_fatture WHERE abs(totale - (imponibile + iva)) >= 0.01",
    )
    assert incoerenti == []


def test_tutte_le_date_nel_mese_simulato(mese) -> None:
    data_dir, _ = mese
    for vista in ("v_fatture", "v_ddt", "v_sal", "v_rapportini"):
        fuori = query(
            data_dir,
            f"SELECT id FROM {vista} WHERE data < DATE '2026-06-01' OR data > DATE '2026-06-30'",
        )
        assert fuori == [], f"{vista} ha documenti fuori dal mese simulato"


def test_tutte_le_entita_validate(mese) -> None:
    data_dir, _ = mese
    for vista in ("v_fatture", "v_ddt", "v_sal", "v_rapportini", "v_pozzetti"):
        bozze = query(data_dir, f"SELECT id FROM {vista} WHERE stato <> 'validato'")
        assert bozze == []


def test_ogni_mutazione_e_committata(mese) -> None:
    """Invariante: nessuna modifica non committata (working tree pulito)."""
    from git import Repo

    data_dir, _ = mese
    assert Repo(data_dir).is_dirty(untracked_files=True) is False


# --------------------------------------------------------- aggregati / viste


def test_totale_fatture_coerente(mese) -> None:
    data_dir, sc = mese
    tot = query(data_dir, "SELECT COALESCE(SUM(totale), 0) AS t FROM v_fatture")[0]["t"]
    assert abs(tot - sc.totale_fatture) < TOLL


def test_ore_e_costo_manodopera_per_cantiere(mese) -> None:
    data_dir, sc = mese
    righe = query(
        data_dir,
        "SELECT cantiere_id, SUM(ore) AS ore, SUM(costo) AS costo "
        "FROM v_rapportini_righe GROUP BY cantiere_id",
    )
    per_cantiere = {r["cantiere_id"]: r for r in righe}
    assert set(per_cantiere) == set(sc.cantieri)
    for cid in sc.cantieri:
        assert abs(per_cantiere[cid]["ore"] - sc.ore[cid]) < TOLL
        assert abs(per_cantiere[cid]["costo"] - sc.costo_mano[cid]) < TOLL


# --------------------------------------------------------- cruscotto / registro


def test_cruscotto_globale(client_mese: TestClient, mese) -> None:
    _data_dir, sc = mese
    admin = _accedi(client_mese, "ufficio01", sc.pin["ufficio01"])
    corpo = client_mese.get("/api/dashboard/costs", headers=admin).json()
    assert corpo["totali"]["n_fatture"] == sc.n_fatture
    assert abs(corpo["totali"]["totale"] - sc.totale_fatture) < TOLL
    assert len(corpo["per_cantiere"]) == 10
    assert len(corpo["cronoprogramma"]) == 10   # un cronoprogramma per cantiere


def test_registro_di_ogni_cantiere(client_mese: TestClient, mese) -> None:
    _data_dir, sc = mese
    admin = _accedi(client_mese, "ufficio02", sc.pin["ufficio02"])
    for cid in sc.cantieri:
        reg = client_mese.get(f"/api/cantieri/{cid}/registro", headers=admin).json()
        assert abs(reg["totali"]["speso_fatture"] - sc.speso[cid]) < TOLL
        assert reg["totali"]["pozzetti"]["totale"] == 4
        assert reg["totali"]["cronoprogramma"]["voci_totali"] == 3
        assert len(reg["sal"]) == 2
        assert reg["totali"]["avanzamento"] == 55.0   # ultimo SAL


def test_scostamenti_consuntivo_abbinato(mese) -> None:
    data_dir, sc = mese
    righe = query(data_dir, "SELECT cantiere_id, consuntivo_abbinato FROM v_cantiere_scostamento")
    per_cantiere = {r["cantiere_id"]: r["consuntivo_abbinato"] for r in righe}
    for cid in sc.cantieri:
        assert abs(per_cantiere[cid] - sc.consuntivo[cid]) < TOLL


# --------------------------------------------------------- report Excel


def test_report_excel_completo(client_mese: TestClient, mese) -> None:
    _data_dir, sc = mese
    admin = _accedi(client_mese, "ufficio03", sc.pin["ufficio03"])
    r = client_mese.get("/api/reports/mensile.xlsx", headers=admin)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    for foglio in ("Riepilogo", "Fatture", "DDT", "Ore", "SAL",
                   "Scostamento computo", "Cronoprogramma", "Pozzetti"):
        assert foglio in wb.sheetnames
    # intestazione + 50 fatture + riga TOTALE
    assert wb["Fatture"].max_row == 1 + sc.n_fatture + 1
    # filtro per cantiere: solo le fatture di CNT-001
    r1 = client_mese.get("/api/reports/mensile.xlsx?cantiere_id=CNT-001", headers=admin)
    ws = load_workbook(io.BytesIO(r1.content))["Fatture"]
    cantieri = {ws.cell(row=riga, column=5).value for riga in range(2, ws.max_row)}
    assert cantieri <= {"CNT-001"}


# --------------------------------------------------------- RBAC (ufficio vs cantiere)


def test_rbac_capocantiere_vede_solo_il_suo(client_mese: TestClient, mese) -> None:
    _data_dir, sc = mese
    capo = client_mese.post(
        "/api/auth/login", json={"username": "capo01", "pin": sc.pin["capo01"]}
    ).json()
    assert [c["id"] for c in capo["utente"]["cantieri"]] == ["CNT-001"]
    assert capo["utente"]["ruolo"] == "operatore"
    intest = {"Authorization": f"Bearer {capo['token']}"}
    # cruscotto e registro sono dell'ufficio (admin): all'operatore sono negati
    assert client_mese.get("/api/dashboard/costs", headers=intest).status_code == 403
    assert client_mese.get("/api/cantieri/CNT-002/registro", headers=intest).status_code == 403


def test_rbac_ufficio_vede_tutti_i_cantieri(client_mese: TestClient, mese) -> None:
    _data_dir, sc = mese
    admin = client_mese.post(
        "/api/auth/login", json={"username": "ufficio01", "pin": sc.pin["ufficio01"]}
    ).json()
    assert admin["utente"]["ruolo"] == "admin"
    assert len(admin["utente"]["cantieri"]) == 10   # l'ufficio vede tutto


def test_giorni_lavorativi_del_mese() -> None:
    """Sanity: giugno 2026 ha 22 giorni lavorativi (base del volume rapportini)."""
    giorni = [
        d for d in (dt.date(2026, 6, 1) + dt.timedelta(days=i) for i in range(30))
        if d.month == 6 and d.weekday() < 5
    ]
    assert len(giorni) == 22
