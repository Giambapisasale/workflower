"""M21 — registri e riepiloghi automatici: pozzetti e cronoprogramma.

Registri che si popolano dalle entità e si aggiornano da soli a ogni documento
(§3.8). Il cronoprogramma confronta pianificato vs consuntivo (dal SAL); i
pozzetti sono un registro con stato. Esposti in cruscotto, registro di cantiere
e nel report Excel (fogli dedicati).
"""

import io

from aiuti import accedi
from openpyxl import load_workbook

from app.core.report import genera_report
from app.core.views import query

# ------------------------------------------------------- viste derivate


def test_cronoprogramma_confronta_pianificato_e_reale(dati_rw) -> None:
    riga = query(dati_rw, "SELECT * FROM v_cronoprogramma WHERE cantiere_id = 'CNT-001'")[0]
    assert riga["voci_totali"] == 3
    assert riga["pianificato_pct"] > 0
    assert riga["reale_pct"] == 32.6  # dall'ultimo SAL del cantiere
    # in ritardo: il reale è sotto il pianificato → delta negativo
    assert riga["delta_pct"] == round(riga["reale_pct"] - riga["pianificato_pct"], 1)


def test_pozzetti_riepilogo_per_stato(dati_rw) -> None:
    r = query(dati_rw, "SELECT * FROM v_pozzetti_riepilogo WHERE cantiere_id = 'CNT-001'")[0]
    assert r["totale"] == 3
    assert (r["previsti"], r["installati"], r["collaudati"]) == (1, 1, 1)


def test_cronoprogramma_si_aggiorna_col_nuovo_sal(client, dati_rw) -> None:
    """Aggiungere un SAL più recente sposta il consuntivo del cronoprogramma."""
    admin = accedi(client, "giovanna")
    prima = query(dati_rw, "SELECT reale_pct FROM v_cronoprogramma WHERE cantiere_id='CNT-001'")
    assert prima[0]["reale_pct"] == 32.6
    # nuovo SAL più recente al 50%
    r = client.post(
        "/api/entities/sal",
        headers=admin,
        json={"dati": {"cantiere_id": "CNT-001", "numero": "4", "data": "2026-07-15",
                       "importo_lavori": 1720000.0, "importo_progressivo": 860000.0,
                       "percentuale_avanzamento": 50.0}},
    )
    assert r.status_code == 200, r.text
    dopo = query(dati_rw, "SELECT reale_pct FROM v_cronoprogramma WHERE cantiere_id='CNT-001'")
    assert dopo[0]["reale_pct"] == 50.0  # la vista si è aggiornata da sola


# ------------------------------------------------------- registro / cruscotto


def test_registro_cantiere_include_registri(client) -> None:
    admin = accedi(client, "giovanna")
    reg = client.get("/api/cantieri/CNT-001/registro", headers=admin).json()
    assert reg["totali"]["cronoprogramma"]["voci_totali"] == 3
    assert reg["totali"]["pozzetti"]["totale"] == 3
    assert len(reg["pozzetti"]) == 3
    assert len(reg["cronoprogramma"]) == 3


def test_dashboard_espone_cronoprogramma(client) -> None:
    admin = accedi(client, "giovanna")
    corpo = client.get("/api/dashboard/costs", headers=admin).json()
    crono = {r["cantiere_id"]: r for r in corpo["cronoprogramma"]}
    assert "CNT-001" in crono
    assert crono["CNT-001"]["reale_pct"] == 32.6


# ------------------------------------------------------- report xlsx


def test_report_xlsx_ha_fogli_dedicati(dati_rw) -> None:
    contenuto = genera_report(dati_rw)
    wb = load_workbook(io.BytesIO(contenuto))
    assert "Cronoprogramma" in wb.sheetnames
    assert "Pozzetti" in wb.sheetnames
    # il foglio pozzetti ha intestazione + 3 righe di CNT-001
    ws = wb["Pozzetti"]
    assert ws.max_row >= 4
