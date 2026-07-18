"""M11 — report Excel: generazione deterministica dalle viste + endpoint."""

from io import BytesIO
from pathlib import Path

from aiuti import accedi
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.report import genera_report

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_report_ha_tutti_i_fogli(dati_rw: Path) -> None:
    wb = load_workbook(BytesIO(genera_report(dati_rw)))
    assert wb.sheetnames == ["Riepilogo", "Fatture", "DDT", "Ore", "SAL", "Scostamento computo"]


def test_report_fatture_con_totale_e_formato_euro(dati_rw: Path) -> None:
    wb = load_workbook(BytesIO(genera_report(dati_rw)))
    ws = wb["Fatture"]
    assert [c.value for c in ws[1]][:3] == ["ID", "Numero", "Data"]
    # 5 fatture del seed + riga TOTALE
    assert ws.max_row == 1 + 5 + 1
    assert ws.cell(row=ws.max_row, column=1).value == "TOTALE"
    # la somma dei totali coincide con quella del seed
    totale = ws.cell(row=ws.max_row, column=8).value
    assert round(totale, 2) == round(10162.60 + 6289.80 + 8344.80 + 4880.00 + 2881.64, 2)
    assert "€" in ws.cell(row=2, column=8).number_format


def test_report_scostamento_dal_computo(dati_rw: Path) -> None:
    wb = load_workbook(BytesIO(genera_report(dati_rw)))
    ws = wb["Scostamento computo"]
    intest = [c.value for c in ws[1]]
    assert "Consuntivo" in intest and "Consumo" in intest
    # 6 voci CNT-001 + 4 voci CNT-002 + TOTALE
    assert ws.max_row == 1 + 10 + 1


def test_report_filtrato_per_cantiere(dati_rw: Path) -> None:
    wb = load_workbook(BytesIO(genera_report(dati_rw, cantiere_id="CNT-002")))
    ws = wb["Fatture"]
    cantieri = {ws.cell(row=r, column=5).value for r in range(2, ws.max_row)}  # esclude TOTALE
    assert cantieri == {"CNT-002"}


def test_endpoint_scarica_xlsx(client: TestClient) -> None:
    intestazioni = accedi(client, "giovanna")
    risposta = client.get("/api/reports/mensile.xlsx?cantiere_id=CNT-001", headers=intestazioni)
    assert risposta.status_code == 200
    assert risposta.headers["content-type"] == XLSX_MIME
    assert ".xlsx" in risposta.headers["content-disposition"]
    wb = load_workbook(BytesIO(risposta.content))
    assert "Fatture" in wb.sheetnames


def test_endpoint_report_riservato_admin(client: TestClient) -> None:
    operatore = accedi(client, "salvo")
    assert client.get("/api/reports/mensile.xlsx", headers=operatore).status_code == 403
